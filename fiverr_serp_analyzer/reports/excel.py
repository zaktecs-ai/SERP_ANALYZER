"""Excel report generation for the Fiverr SERP Analyzer.

Produces fiverr_serp_analysis.xlsx with multiple sheets, formatting,
filters, freeze panes, and conditional formatting on score columns.
"""

import os
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule, ColorScaleRule


# Styling constants
HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)
CELL_ALIGNMENT = Alignment(vertical="top", wrap_text=False)
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)
SCORE_FILL_HIGH = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
SCORE_FILL_MED = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
SCORE_FILL_LOW = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")


def _style_header(ws, headers: list, row: int = 1):
    """Apply header styling to a row."""
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT
        cell.border = THIN_BORDER


def _style_data_cell(ws, row: int, col: int, value):
    """Apply data cell styling."""
    cell = ws.cell(row=row, column=col, value=value)
    cell.alignment = CELL_ALIGNMENT
    cell.border = THIN_BORDER
    return cell


def _auto_width(ws, min_width: int = 8, max_width: int = 40):
    """Auto-fit column widths."""
    for col_cells in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        adjusted = min(max(max_len + 2, min_width), max_width)
        ws.column_dimensions[col_letter].width = adjusted


def _add_score_conditional_format(ws, col_letter: str, start_row: int, end_row: int):
    """Add conditional formatting for score columns (0-100)."""
    cell_range = f"{col_letter}{start_row}:{col_letter}{end_row}"
    ws.conditional_formatting.add(
        cell_range,
        ColorScaleRule(
            start_type="num", start_value=0, start_color="FFC7CE",
            mid_type="num", mid_value=50, mid_color="FFEB9C",
            end_type="num", end_value=100, end_color="C6EFCE",
        ),
    )


def generate_excel(
    keyword_analyses: list,
    all_gigs: list,
    clusters: dict,
    errors: list,
    run_info: dict,
    output_path: str = "fiverr_serp_analysis.xlsx",
):
    """Generate the full Excel report with all sheets.

    Args:
        keyword_analyses: List of dicts with per-keyword analysis.
        all_gigs: List of dicts with all gig data.
        clusters: Dict of keyword clusters.
        errors: List of error dicts.
        run_info: Dict with collection run metadata.
        output_path: Output file path.
    """
    wb = Workbook()

    # Remove default sheet
    wb.remove(wb.active)

    # --- Sheet 1: Keyword Summary ---
    ws1 = wb.create_sheet("Keyword Summary")
    headers1 = [
        "Keyword", "Intent", "Intent Score", "Demand Signal",
        "Total Results", "Gigs Collected", "Median Reviews",
        "Median Rating", "Competition Score", "SERP Strength Score",
        "Relevance Score", "Opportunity Score", "Cluster", "Status",
    ]
    _style_header(ws1, headers1)

    for row_idx, analysis in enumerate(keyword_analyses, 2):
        _style_data_cell(ws1, row_idx, 1, analysis.get("keyword"))
        _style_data_cell(ws1, row_idx, 2, analysis.get("primary_intent"))
        _style_data_cell(ws1, row_idx, 3, analysis.get("intent_score"))
        _style_data_cell(ws1, row_idx, 4, analysis.get("demand_signal"))
        _style_data_cell(ws1, row_idx, 5, analysis.get("total_results"))
        _style_data_cell(ws1, row_idx, 6, analysis.get("gig_count"))
        _style_data_cell(ws1, row_idx, 7, analysis.get("median_reviews"))
        _style_data_cell(ws1, row_idx, 8, analysis.get("median_rating"))
        _style_data_cell(ws1, row_idx, 9, analysis.get("competition_score"))
        _style_data_cell(ws1, row_idx, 10, analysis.get("serp_strength_score"))
        _style_data_cell(ws1, row_idx, 11, analysis.get("relevance_score"))
        _style_data_cell(ws1, row_idx, 12, analysis.get("opportunity_score"))
        _style_data_cell(ws1, row_idx, 13, analysis.get("cluster_name", ""))
        _style_data_cell(ws1, row_idx, 14, analysis.get("status", ""))

    _auto_width(ws1)
    ws1.freeze_panes = "A2"
    ws1.auto_filter.ref = ws1.dimensions

    # Conditional formatting on score columns
    last_row = len(keyword_analyses) + 1
    if last_row > 1:
        for col in ["C", "I", "J", "K", "L"]:
            _add_score_conditional_format(ws1, col, 2, last_row)

    # --- Sheet 2: Top 20 Gigs ---
    ws2 = wb.create_sheet("Top 20 Gigs")
    headers2 = [
        "Keyword", "Position", "Title", "URL", "Seller", "Seller Level",
        "Rating", "Reviews", "Starting Price", "Delivery Time",
        "Category", "Badges", "Tags",
    ]
    _style_header(ws2, headers2)

    for row_idx, gig in enumerate(all_gigs, 2):
        _style_data_cell(ws2, row_idx, 1, gig.get("keyword"))
        _style_data_cell(ws2, row_idx, 2, gig.get("serp_position"))
        _style_data_cell(ws2, row_idx, 3, gig.get("title_normalized"))
        _style_data_cell(ws2, row_idx, 4, gig.get("url_normalized"))
        _style_data_cell(ws2, row_idx, 5, gig.get("seller_name_normalized"))
        _style_data_cell(ws2, row_idx, 6, gig.get("seller_level_normalized"))
        _style_data_cell(ws2, row_idx, 7, gig.get("seller_rating_normalized"))
        _style_data_cell(ws2, row_idx, 8, gig.get("review_count_cleaned"))
        _style_data_cell(ws2, row_idx, 9, gig.get("starting_price_normalized"))
        _style_data_cell(ws2, row_idx, 10, gig.get("delivery_time_normalized"))
        _style_data_cell(ws2, row_idx, 11, gig.get("category_normalized"))
        _style_data_cell(ws2, row_idx, 12,
                         ", ".join(gig.get("badges_normalized", [])) if gig.get("badges_normalized") else "")
        _style_data_cell(ws2, row_idx, 13,
                         ", ".join(gig.get("service_tags_normalized", [])) if gig.get("service_tags_normalized") else "")

    _auto_width(ws2)
    ws2.freeze_panes = "A2"
    ws2.auto_filter.ref = ws2.dimensions

    # --- Sheet 3: Competitor Analysis ---
    ws3 = wb.create_sheet("Competitor Analysis")
    headers3 = [
        "Keyword", "Position", "Title", "Title Opt Score",
        "Keyword Relevance", "Seller Strength", "Review Strength",
        "Rating Strength", "Price Positioning",
    ]
    _style_header(ws3, headers3)

    row_idx = 2
    for analysis in keyword_analyses:
        gig_scores = analysis.get("gig_scores", [])
        for gs in gig_scores:
            _style_data_cell(ws3, row_idx, 1, analysis.get("keyword"))
            _style_data_cell(ws3, row_idx, 2, gs.get("serp_position"))
            _style_data_cell(ws3, row_idx, 3, gs.get("title", "")[:100])
            _style_data_cell(ws3, row_idx, 4, gs.get("title_optimization_score"))
            _style_data_cell(ws3, row_idx, 5, gs.get("keyword_relevance_score"))
            _style_data_cell(ws3, row_idx, 6, gs.get("seller_strength_score"))
            _style_data_cell(ws3, row_idx, 7, gs.get("review_strength_score"))
            _style_data_cell(ws3, row_idx, 8, gs.get("rating_strength_score"))
            _style_data_cell(ws3, row_idx, 9, gs.get("price_positioning_score"))
            row_idx += 1

    _auto_width(ws3)
    ws3.freeze_panes = "A2"
    ws3.auto_filter.ref = ws3.dimensions

    # --- Sheet 4: Keyword Metrics ---
    ws4 = wb.create_sheet("Keyword Metrics")
    headers4 = [
        "Keyword", "Exact Title Match %", "Partial Title Match %",
        "Token Match %", "Avg Token Ratio", "Clearly Offering %",
        "Established Seller %", "Top 5 Review Share %",
        "Top 10 Review Share %", "Top 5 vs Bottom 15 Ratio",
    ]
    _style_header(ws4, headers4)

    for row_idx, analysis in enumerate(keyword_analyses, 2):
        _style_data_cell(ws4, row_idx, 1, analysis.get("keyword"))
        _style_data_cell(ws4, row_idx, 2, analysis.get("exact_title_match_pct"))
        _style_data_cell(ws4, row_idx, 3, analysis.get("partial_title_match_pct"))
        _style_data_cell(ws4, row_idx, 4, analysis.get("token_match_pct"))
        _style_data_cell(ws4, row_idx, 5, analysis.get("avg_token_match_ratio"))
        _style_data_cell(ws4, row_idx, 6, analysis.get("clearly_offering_pct"))
        _style_data_cell(ws4, row_idx, 7, analysis.get("established_seller_concentration"))
        _style_data_cell(ws4, row_idx, 8, analysis.get("top5_review_share"))
        _style_data_cell(ws4, row_idx, 9, analysis.get("top10_review_share"))
        _style_data_cell(ws4, row_idx, 10, analysis.get("top5_vs_bottom15_strength"))

    _auto_width(ws4)
    ws4.freeze_panes = "A2"
    ws4.auto_filter.ref = ws4.dimensions

    # --- Sheet 5: Opportunity Ranking ---
    ws5 = wb.create_sheet("Opportunity Ranking")
    headers5 = [
        "Rank", "Keyword", "Opportunity Score", "Demand Score",
        "Intent Score", "Relevance Score", "Competition Score",
        "SERP Weakness Score", "Intent Type", "Demand Signal",
    ]
    _style_header(ws5, headers5)

    # Sort by opportunity score descending
    ranked = sorted(keyword_analyses, key=lambda x: x.get("opportunity_score", 0), reverse=True)
    for row_idx, analysis in enumerate(ranked, 2):
        _style_data_cell(ws5, row_idx, 1, row_idx - 1)
        _style_data_cell(ws5, row_idx, 2, analysis.get("keyword"))
        _style_data_cell(ws5, row_idx, 3, analysis.get("opportunity_score"))
        _style_data_cell(ws5, row_idx, 4, analysis.get("demand_score"))
        _style_data_cell(ws5, row_idx, 5, analysis.get("intent_score"))
        _style_data_cell(ws5, row_idx, 6, analysis.get("relevance_score"))
        _style_data_cell(ws5, row_idx, 7, analysis.get("competition_score"))
        _style_data_cell(ws5, row_idx, 8, analysis.get("serp_weakness_score"))
        _style_data_cell(ws5, row_idx, 9, analysis.get("primary_intent"))
        _style_data_cell(ws5, row_idx, 10, analysis.get("demand_signal"))

    _auto_width(ws5)
    ws5.freeze_panes = "A2"
    ws5.auto_filter.ref = ws5.dimensions

    last_row5 = len(ranked) + 1
    if last_row5 > 1:
        for col in ["C", "D", "E", "F", "G", "H"]:
            _add_score_conditional_format(ws5, col, 2, last_row5)

    # --- Sheet 6: Keyword Clusters ---
    ws6 = wb.create_sheet("Keyword Clusters")
    headers6 = ["Cluster", "Cluster Name", "Keyword Count", "Keywords"]
    _style_header(ws6, headers6)

    row_idx = 2
    for cid, cdata in clusters.items():
        _style_data_cell(ws6, row_idx, 1, cid)
        _style_data_cell(ws6, row_idx, 2, cdata.get("name", ""))
        _style_data_cell(ws6, row_idx, 3, cdata.get("size", 0))
        _style_data_cell(ws6, row_idx, 4, ", ".join(cdata.get("keywords", [])))
        row_idx += 1

    _auto_width(ws6, max_width=80)
    ws6.freeze_panes = "A2"

    # --- Sheet 7: Methodology ---
    ws7 = wb.create_sheet("Methodology")
    methodology_text = [
        ["Fiverr SERP Analyzer — Scoring Methodology"],
        [""],
        ["Opportunity Score Formula:"],
        ["  Opportunity = Demand*0.25 + Intent*0.20 + Relevance*0.20"],
        ["              + (100-Competition)*0.25 + (100-SERP_Strength)*0.10"],
        [""],
        ["Component Definitions:"],
        ["  Demand Score: Based on Fiverr's visible total result count (NOT search volume)."],
        ["    <100 results = 10, <500 = 25, <2000 = 50, <10000 = 75, >=10000 = 100"],
        ["  Intent Score: Rule-based NLP classification of buyer intent (0-100)."],
        ["    Transactional/service-specific keywords score highest."],
        ["  Relevance Score: How closely top-20 gig titles match the keyword."],
        ["    Exact match %, partial match %, token match ratio."],
        ["  Competition Score: Multi-signal (median reviews, median rating,"],
        ["    established seller concentration, review distribution)."],
        ["  SERP Strength Score: Concentration of reviews among top sellers."],
        ["    Higher concentration = stronger SERP = lower opportunity."],
        [""],
        ["All scores are 0-100. Higher Opportunity = better keyword to target."],
        ["Missing data is null, never zero. Missing reviews != 0 reviews."],
        ["Missing price != $0. All values labeled: observed | calculated | assumption."],
        [""],
        ["Data Quality States: extracted | missing | inferred | parsing_failed"],
        ["Every extracted field carries its state and the selector that matched."],
    ]
    for row_idx, row_data in enumerate(methodology_text, 1):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws7.cell(row=row_idx, column=col_idx, value=value)
            if row_idx == 1:
                cell.font = Font(bold=True, size=14)
    ws7.column_dimensions["A"].width = 100

    # --- Sheet 8: Errors ---
    ws8 = wb.create_sheet("Errors")
    headers8 = ["Keyword", "Error Type", "Error Detail", "Timestamp"]
    _style_header(ws8, headers8)

    for row_idx, err in enumerate(errors, 2):
        _style_data_cell(ws8, row_idx, 1, err.get("keyword", ""))
        _style_data_cell(ws8, row_idx, 2, err.get("error_type", ""))
        _style_data_cell(ws8, row_idx, 3, err.get("error_detail", ""))
        _style_data_cell(ws8, row_idx, 4, err.get("timestamp", ""))

    _auto_width(ws8)
    ws8.freeze_panes = "A2"

    # --- Sheet 9: Collection Runs ---
    ws9 = wb.create_sheet("Collection Runs")
    headers9 = ["Run ID", "Started", "Finished", "Total Keywords",
                "Completed", "Failed", "Status"]
    _style_header(ws9, headers9)

    if run_info:
        _style_data_cell(ws9, 2, 1, run_info.get("run_id", ""))
        _style_data_cell(ws9, 2, 2, run_info.get("started_at", ""))
        _style_data_cell(ws9, 2, 3, run_info.get("finished_at", ""))
        _style_data_cell(ws9, 2, 4, run_info.get("total_keywords", 0))
        _style_data_cell(ws9, 2, 5, run_info.get("completed_keywords", 0))
        _style_data_cell(ws9, 2, 6, run_info.get("failed_keywords", 0))
        _style_data_cell(ws9, 2, 7, run_info.get("status", ""))

    _auto_width(ws9)
    ws9.freeze_panes = "A2"

    # Save
    wb.save(output_path)
    print(f"Excel report saved to: {output_path}")
    return output_path